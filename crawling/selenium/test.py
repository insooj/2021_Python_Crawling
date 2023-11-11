from typing import KeysView, List
from urllib import request
from bs4 import BeautifulSoup, element
import urllib.request
import urllib.parse
from openpyxl.workbook import Workbook
from selenium import webdriver
from openpyxl import workbook
import openpyxl
import time
import selenium
import requests as rq
from selenium.webdriver.chrome import options
from selenium.webdriver.common.keys import Keys 
from urllib.request import urlopen
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotVisibleException
import re
import numpy
import gettext
import gzip
import tkinter as tk
from tkinter import *
import tkinter.font
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
# import requests
# import pandas as pd
# import numpy as np
# import folium
# from folium.plugins import MiniMap
# import webview

#------------------------------------------------------------------------------------------------------------------------
# usb 에러 코드 해결
# options=webdriver.ChromeOptions()
# options.add_experimental_option("excludeSwitches", ["enable-logging"])
# driver = webdriver.Chrome(options=options)


# driver.maximize_window()    
# url = 'https://www.kda.or.kr/'
# driver.get(url)

# time.sleep(2)
# driver.find_element_by_css_selector('.m91').click()

# driver.find_element_by_link_text('우리동네 치과의사 실명제 치과찾기').click()

# time.sleep(2)
# html = driver.page_source
# soup = BeautifulSoup(html,'html.parser')
# rank = soup.select('#dentalFindListArea > div.table-list.line > table > tbody > tr:nth-child('+str(1)+') > td:nth-child(1)')

# numbers = re.sub(r'[^0-9]', '', str(rank))


# selArray=[['a' for _ in range(5)]for _ in range(int(numbers))]
# i=0

# selWb=openpyxl.Workbook()
# selws = selWb.active
# selws = selWb.create_sheet('실명제 치과 데이터')
# selws.append(['순위','치과명','원장 이름','전화번호','주소'])
# for Screen in range(1,24):
#     for page in range(3,8):
#         html = driver.page_source
#         soup = BeautifulSoup(html,'html.parser')
#         for list in range(1,16):  
# #        #각 카테고리별로 크롤링해주기   
               
#             rank=soup.select('#dentalFindListArea > div.table-list.line > table > tbody > tr:nth-child('+str(list)+') > td:nth-child(1)')
#             name=soup.select('#dentalFindListArea > div.table-list.line > table > tbody > tr:nth-child('+str(list)+') > td:nth-child(2) > a')
#             name2=soup.select('#dentalFindListArea > div.table-list.line > table > tbody > tr:nth-child('+str(list)+') > td:nth-child(3)')
#             num=soup.select('#dentalFindListArea > div.table-list.line > table > tbody > tr:nth-child('+str(list)+') > td:nth-child(4)')
#             ad=soup.select('#dentalFindListArea > div.table-list.line > table > tbody > tr:nth-child('+str(list)+') > td.txt_ac.tL')
#             all=zip(rank,name,name2,num,ad)
            
# #       #리스트로 분리
#             for a,b,c,d,e in all:
#                 selArray[i][0] = a.text
#                 selArray[i][1] = b.text
#                 selArray[i][2] = c.text
#                 selArray[i][3] = d.text
#                 selArray[i][4] = e.text
#                 print('순위:',selArray[i][0],"*",'치과명:',selArray[i][1],"*",'원장 이름:',selArray[i][2],"*",'전화번호:',selArray[i][3],"*",'주소:',selArray[i][4],"*") 
#                 i=i+1
#             selws.append([a.text,b.text,c.text,d.text,e.text])    
                
                
            
#         list = 0    
#         if((a.text)=="1"):
#             break
#         # 페이지 넘기기
#         driver.find_element_by_xpath('//*[@id="dentalFindListArea"]/div[5]/a[%s+1]' %page).click()
             

# selWb.save("치과 데이터.xlsx")
#크롤링 데이터 수집, 저장------------------------------------------------------------------------------------------------------------------------
# x = input("지역을 입력하시오>>( OO시 OO구 OO동/OO로 ) ")

# address = x
# cnt =0
# for radd in range(1, int(numbers)):
#     if(address in selArray[radd][4]):
#         cnt = cnt +1

# dentIndex =[['a' for _ in range(2)]for _ in range(int(cnt))]

# i=0
# for radd in range(1, int(numbers)):
#     if(address in selArray[radd][4]):
#         dentIndex[i][0] = selArray[radd][1]
#         dentIndex[i][1] = selArray[radd][4]
#         print(i, '번', '치과명:',dentIndex[i][0],'주소:',dentIndex[i][1])
#         i = i + 1
        
        
# x1 = input("해당 치과의 번호를 입력해주세요>>") 
# x2 = int(x1)
# seletedentistry = ''
# for d in range(0,1):
#     seletedentistry = dentIndex[(x2)][0] 
#------------------------------------------------------------------------------------------------------------------------

# # 선택한 치과의 네이버 리뷰창으로 이동

# baseUrl = 'https://map.naver.com/v5/search/'
# plusurl = seletedentistry
# url = baseUrl + seletedentistry
# driver.get(url)
# try: 
#     driver.switch_to.frame('searchIframe')
#     x = input("지역: ")
#     html = driver.page_source
#     soup = BeautifulSoup(html,'html.parser')
#     for i in range(1,10):
    
#         seladdress = driver.find_element_by_xpath('//*[@id="_pcmap_list_scroll_container"]/ul/li['+str(i)+']/div[2]/div/div/span/a/span[1]').text
#         if( x in seladdress):
#             driver.find_element_by_xpath('//*[@id="_pcmap_list_scroll_container"]/ul/li['+str(i)+']/div[2]/a[1]/div/div/span[1]').click()
#     driver.switch_to.default_content()
#     driver.implicitly_wait(3)
#     driver.switch_to.frame('entryIframe')

#     driver.find_element_by_xpath('//*[@id="app-root"]/div/div/div[2]/div[3]/div/div/div/div/a[2]').click()

# except selenium.common.exceptions.ElementClickInterceptedException as e:
#     driver.switch_to.default_content()
#     driver.implicitly_wait(3)
#     driver.switch_to.frame('entryIframe')

#     driver.find_element_by_xpath('//*[@id="app-root"]/div/div/div[2]/div[3]/div/div/div/div/a[2]').click()



# #  첫번째 리뷰 창에 있는 리뷰들을 모두 출력
# html = driver.page_source
# soup = BeautifulSoup(html,'html.parser')
# print("=====================================")
# time.sleep(1)
# try:
#     wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section._35EJ4 > div._2oZg_ > span._1fvo3.Sv1wj > em')
# except selenium.common.exceptions.NoSuchElementException as e:
#     wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div > div.place_section._35EJ4 > div._2oZg_ > span._1fvo3.Sv1wj > em')

# print(wjatn.text)

# try:
#     Wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section._35EJ4 > div._2oZg_ > span:nth-child(2)')
# except selenium.common.exceptions.NoSuchElementException as e:
#     Wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div > div.place_section._35EJ4 > div._2oZg_ > span:nth-child(2)')

# print(Wjatn.text)
# while True:
#         try:
#             time.sleep(0.3)
#             driver.find_element_by_css_selector("a._3iTUo").click()
            
#         except selenium.common.exceptions.NoSuchElementException as e:
#             break
# time.sleep(0.5)
# review = driver.find_elements_by_css_selector(".WoYOw")
# for i in review:
#     print(i.text)

# driver.close()

class Main(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title("바른치과 길잡이")
        self.geometry("1375x850+100+100")
        self.resizable(0,0)
        
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        def close():
            self.quit()
            self.destroy()

        menubar = tk.Menu(self)

        menu_1=tk.Menu(menubar, tearoff=0)
        menu_1.add_command(label="제작자")
        menu_1.add_command(label="버전")
        menu_1.add_separator()
        menu_1.add_command(label="종료하기", command=close)
        menubar.add_cascade(label="메뉴", menu=menu_1)

        self.config(menu=menubar)
        
        self.frames = {}
        for F in (StartPage, PageOne):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("StartPage")

    def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()

class StartPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        buttonStart = tk.Button(self, overrelief="solid", text="시작하기", command=lambda: controller.show_frame("PageOne"))

        font = tk.font.Font(family="맑은고딕", size=50)

        labelTitle = tk.Label(self,text="바른치과 길잡이", font=font)
        labelMake = tk.Label(self, text="Version_0.0\nmade by 서준일, 신준호, 장인수", justify = "right")

        labelTitle.place(relx = 0.35, rely = 0.3)
        labelMake.pack(side="right", anchor="s")
        buttonStart.place(relx = 0.45, rely = 0.7, width = 100, height = 50)


class PageOne(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        options=webdriver.ChromeOptions()
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        driver = webdriver.Chrome(options=options)


        SelectDentArray=[['a' for _ in range(500)]for _ in range(int(2))]
        SelcetDentCount = 0
        global entrytext
        def getAdress():
            global entrytext
            entrytext = entry.get()
            self.address = entry.get()
            showListbox()
            

        def showListbox():        
            SelcetDentCount = 0
            cnt = 0
            listbox.delete(0, END)
            for item in col_address:
                if(entrytext in item.value):
                    SelectDentArray[0][SelcetDentCount] = DentArray[0][cnt]
                    SelectDentArray[1][SelcetDentCount] = DentArray[3][cnt]

                    strSelectDent = '치과이름: ' + SelectDentArray[0][SelcetDentCount] +'  주소: ' + SelectDentArray[1][SelcetDentCount]

                    SelcetDentCount = SelcetDentCount + 1

                    listbox.insert(END, strSelectDent)
                cnt = cnt+1

        def reviewCrawling():
            for item in listbox.curselection():
                item
            baseUrl = 'https://map.naver.com/v5/search/'
            plusurl = SelectDentArray[0][item]
            url = baseUrl + SelectDentArray[0][item]
            
            driver.get(url)
            html = driver.page_source
            soup = BeautifulSoup(html,'html.parser')
            time.sleep(1)
            try: 
                driver.switch_to.frame('searchIframe')
                x = entrytext
                html = driver.page_source
                soup = BeautifulSoup(html,'html.parser')
                seladdresses = driver.find_elements_by_css_selector("span._3hCbH")
                for j in seladdresses:
                    i = j.text
                   
                    if( x in i):
                        driver.implicitly_wait(3)
                        driver.find_element_by_class_name("_2w9xx").click()
                        # driver.find_element_by_xpath('//*[@id="_pcmap_list_scroll_container"]/ul/li['+str(i)+']/div[2]/a[1]/div/div/span[1]').click()

                # for i in range(1,10):
                #     try:
                #         seladdress = driver.find_element_by_xpath('//*[@id="_pcmap_list_scroll_container"]/ul/li['+str(i)+']/div[2]/div/div/span/a/span[1]').text
                #     except selenium.common.exceptions.NoSuchElementException as e: 
                #         saladdress = driver.find_element_by_xpath('//*[@id="_pcmap_list_scroll_container"]/ul/li['+str(i)+']/div[1]/div/div/span/a/span[1]').text
                    
                #     if( x in seladdress):
                #         driver.implicitly_wait(3)
                #         driver.find_element_by_xpath('//*[@id="_pcmap_list_scroll_container"]/ul/li['+str(i)+']/div[2]/a[1]/div/div/span[1]').click()
                                
                driver.switch_to.default_content()
                driver.implicitly_wait(3)
                
                driver.switch_to.frame('entryIframe')

                driver.find_element_by_xpath('//*[@id="app-root"]/div/div/div[2]/div[3]/div/div/div/div/a[2]').click()

            except selenium.common.exceptions.ElementClickInterceptedException as e:
                driver.switch_to.default_content()
                driver.implicitly_wait(3)
                driver.switch_to.frame('entryIframe')

                driver.find_element_by_xpath('//*[@id="app-root"]/div/div/div[2]/div[3]/div/div/div/div/a[2]').click()



            #  첫번째 리뷰 창에 있는 리뷰들을 모두 출력
            html = driver.page_source
            soup = BeautifulSoup(html,'html.parser')
            print("=====================================")
            time.sleep(1)
          
            try:
                wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section._35EJ4 > div._2oZg_ > span._1fvo3.Sv1wj > em')
            except selenium.common.exceptions.NoSuchElementException as e:
                wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div > div.place_section._35EJ4 > div._2oZg_ > span._1fvo3.Sv1wj > em')
            finally : 
                wjatn = None
            try:
                wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section._35EJ4 > div._2oZg_ > span._1fvo3.Sv1wj > em')
            except selenium.common.exceptions.NoSuchElementException as e:
                wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(4) > div:nth-child(4) > div.place_section._35EJ4 > div._2oZg_ > span._1fvo3.Sv1wj > em')
            # #평점출력 -> 레이블 만들어서 거기 텍스트값으로 넣어주고
            #print(wjatn.text)

            try:
                Wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section._35EJ4 > div._2oZg_ > span:nth-child(2)')
            except selenium.common.exceptions.NoSuchElementException as e:
                Wjatn = driver.find_element_by_css_selector('#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div > div.place_section._35EJ4 > div._2oZg_ > span:nth-child(2)')

            #평점 개수
            #print(Wjatn.text)
            while True:
                    try:
                        time.sleep(0.3)
                        driver.find_element_by_css_selector("a._3iTUo").click()

                    except selenium.common.exceptions.NoSuchElementException as e:
                        break
            time.sleep(0.5)
            review = driver.find_elements_by_css_selector(".WoYOw")
            
            #리뷰 출력
            reviewListbox.insert(0,"평점 :★"+ wjatn.text)
            for i in review:
                reviewListbox.insert(END, i.text)




        # def showMap():
        #     tk.PanedWindow = webview.create_window('카카오맵','http://127.0.0.1:1030')
        #     webview.start()


        savingExcel = load_workbook('C:/Users/insu4/Desktop/crawling/selenium/치과 데이터.xlsx')
        SheetExcel = savingExcel.active

        col_name = SheetExcel["B"]      #치과명
        col_name2 = SheetExcel["C"]     #원장 이름
        col_number = SheetExcel["D"]    #전화번호
        col_address = SheetExcel["E"]   #주소

        DentArray=[['a' for _ in range(1666)]for _ in range(int(4))] #발표 전에 확인하고 수정
        i = 0
        for item in col_name:
            DentArray[0][i] =item.value
            i = i+1
        i=0

        for item in col_name2:
            DentArray[1][i] = item.value
            i = i+1
        i=0

        for item in col_number:
            DentArray[2][i] = item.value
            i = i+1
        i=0

        for item in col_address:
            DentArray[3][i] = item.value
            i = i+1
        i=0
        
        eLabel = tk.Label(self, text = "주소를 입력하세요")
        entry = tk.Entry(self)
        entry.insert(0, "")
        buttonEnterAdress = tk.Button(self, overrelief="solid", text="입력", command=getAdress)
        buttonlistbox = tk.Button(self, overrelief="solid", text="선택", command=reviewCrawling)
        label = tk.Label(self, text = "주소")
        
        

        listbox = tk.Listbox(self, selectmode="extended")
        reviewListbox = tk.Listbox(self, selectmode = "extended")

        #def getSelect(self):
        #    selection = listbox.curselection()
        #    if(len(selection) == 0):
        #        return

        #    Selectadress = listbox.get(selection[0])
        #    return Selectadress

        #getSelect(self)

        listbox.yview()
        eLabel.place(relx = 0.05, rely = 0.05)
        entry.place(relx=0.05, rely=0.1, width = 200)
        buttonEnterAdress.place(relx = 0.15, rely = 0.15)
        listbox.place(relx = 0.25, rely = 0.1, width = 500, height = 500)
        buttonlistbox.place(relx = 0.2, rely = 0.6)
        label.place(relx=0.05, rely=0.3)
        reviewListbox.place(relx=0.5, rely = 0.1, width = 500, height = 500)

if __name__ == "__main__":
    app = Main()
    app.mainloop()




